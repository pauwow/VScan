import os
import pandas as pd
from datetime import datetime
import secrets
import string
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

def generate_password(length=14):
    alphabet = string.ascii_letters + string.digits
    return ''.join(secrets.choice(alphabet) for _ in range(length))

def _choose_date_col(df):
    if "transaction_datetime" in df.columns:
        return "transaction_datetime"
    if "TransactionDateTime" in df.columns:
        return "TransactionDateTime"
    for col in df.columns:
        if col.lower().startswith("transaction") and "time" in col.lower():
            return col
    return None

def _build_intervals(entity_data, date_col):
    """Return multi-line string of intervals per day."""
    result_lines = []
    entity_data = entity_data.sort_values(date_col)

    for day, day_df in entity_data.groupby(entity_data[date_col].dt.date):
        times = day_df[date_col].dropna().sort_values().tolist()
        if len(times) <= 1:
            continue

        intervals = []
        for i in range(1, len(times)):
            diff = (times[i] - times[i - 1]).total_seconds() / 60
            intervals.append(str(int(diff)))

        first_time = times[0].strftime("%Y-%m-%d %H:%M")
        line = f"{first_time}: " + " > ".join(intervals)
        result_lines.append(line)

    return "\n".join(result_lines) if result_lines else "N/A"

def _build_interval_column(df, date_col, entity_col):
    """Return a list of intervals per row for the RawData sheet."""
    intervals = []
    df_sorted = df.sort_values([entity_col, date_col])
    last_times = {}

    for idx, row in df_sorted.iterrows():
        entity = row[entity_col]
        current_time = row[date_col]
        if pd.isna(current_time):
            intervals.append(None)
            continue

        if entity in last_times:
            diff = (current_time - last_times[entity]).total_seconds() / 60
            intervals.append(diff)
        else:
            intervals.append(None)

        last_times[entity] = current_time

    return intervals

def summarize_entities(df, entity_col, date_col=None, top_n=20, include_intervals=True):
    summaries = []
    if entity_col not in df.columns:
        return pd.DataFrame()

    if date_col is None:
        date_col = _choose_date_col(df)
    if date_col is None or date_col not in df.columns:
        return pd.DataFrame()

    df_loc = df.copy()
    df_loc[date_col] = pd.to_datetime(df_loc[date_col], errors="coerce")
    df_loc["YearMonth"] = df_loc[date_col].dt.to_period("M")

    top_entities = df_loc[entity_col].value_counts().head(top_n).index
    for entity in top_entities:
        entity_data = df_loc[df_loc[entity_col] == entity]

        day_counts = entity_data[date_col].dt.date.value_counts()
        peak_day = day_counts.idxmax() if not day_counts.empty else None
        peak_count = int(day_counts.max()) if not day_counts.empty else 0
        low_day = day_counts.idxmin() if not day_counts.empty else None
        low_count = int(day_counts.min()) if not day_counts.empty else 0

        entity_label = "Card Number" if entity_col == "card_no" else ("Cashier" if entity_col == "cashier" else entity_col)

        summary = {
            entity_label: str(entity),
            "Month": str(entity_data["YearMonth"].iloc[0]),
            "Total Transactions": len(entity_data),
            "First Transaction": entity_data[date_col].min(),
            "Last Transaction": entity_data[date_col].max(),
            "Day with Most Transactions": f"{peak_day} ({peak_count})" if peak_day else "N/A",
            "Day with Fewest Transactions": f"{low_day} ({low_count})" if low_day else "N/A",
        }

        if include_intervals:
            summary["Transaction Intervals"] = _build_intervals(entity_data, date_col)

        if "branch_code" in entity_data.columns:
            summary["Distinct Branches"] = int(entity_data["branch_code"].nunique())
            summary["Branch List"] = ", ".join(entity_data["branch_code"].dropna().astype(str).unique())
        elif "branch_name" in entity_data.columns:
            summary["Distinct Branches"] = int(entity_data["branch_name"].nunique())
            summary["Branch List"] = ", ".join(entity_data["branch_name"].dropna().astype(str).unique())

        if "cashier" in entity_data.columns and entity_col != "cashier":
            summary["Distinct Cashiers"] = int(entity_data["cashier"].nunique())
            summary["Cashier List"] = ", ".join(entity_data["cashier"].dropna().astype(str).unique())

        if "register_no" in entity_data.columns:
            summary["Distinct Registers"] = int(entity_data["register_no"].nunique())
            summary["Register List"] = ", ".join(entity_data["register_no"].dropna().astype(str).unique())

        if entity_col == "cashier" and "card_no" in entity_data.columns:
            summary["Distinct Cards"] = int(entity_data["card_no"].nunique())
            summary["Cards List"] = ", ".join(entity_data["card_no"].dropna().astype(str).unique())

        if "trans_total" in entity_data.columns:
            summary["Sum of Transaction Total"] = float(entity_data["trans_total"].sum())
        elif "transaction_amount" in entity_data.columns:
            summary["Sum of Transaction Total"] = float(entity_data["transaction_amount"].sum())

        if "point_earned" in entity_data.columns:
            summary["Total Points"] = float(entity_data["point_earned"].sum())

        summaries.append(summary)

    if summaries:
        df_sum = pd.DataFrame(summaries)
        df_sum = df_sum.sort_values("Total Transactions", ascending=False).reset_index(drop=True)
        return df_sum
    else:
        return pd.DataFrame()

def encrypt_excel(input_path, desired_output_path, password):
    try:
        import pythoncom
        import win32com.client as win32
        excel = win32.gencache.EnsureDispatch('Excel.Application')
        excel.DisplayAlerts = False
        wb = excel.Workbooks.Open(os.path.abspath(input_path))
        out_abs = os.path.abspath(desired_output_path)
        wb.SaveAs(out_abs, FileFormat=51, Password=password)
        wb.Close(SaveChanges=False)
        excel.Application.Quit()
        return out_abs
    except Exception:
        pass

    try:
        from msoffcrypto.format.ooxml import OOXMLFile
        out_abs = os.path.abspath(desired_output_path)
        with open(input_path, "rb") as f_in:
            ooxml = OOXMLFile(f_in)
            with open(out_abs, "wb") as f_out:
                ooxml.encrypt(password, f_out)
        return out_abs
    except Exception:
        pass

    try:
        import pyAesCrypt
        bufferSize = 64 * 1024
        if not desired_output_path.lower().endswith(".aes"):
            out_aes = desired_output_path + ".aes"
        else:
            out_aes = desired_output_path
        pyAesCrypt.encryptFile(input_path, out_aes, password, bufferSize)
        return os.path.abspath(out_aes)
    except Exception as e_aes:
        raise RuntimeError(f"Encryption failed with all methods: {e_aes}")

def process_dynamic_schema(df, output_file, top_n_cards=20, top_n_cashiers=20, separate_cards=False, include_intervals=True):
    if "card_no" in df.columns:
        df["card_no"] = df["card_no"].astype(str)

    date_col = _choose_date_col(df)

    # Add interval_minutes column if requested
    if include_intervals and date_col and "card_no" in df.columns:
        df["interval_minutes"] = _build_interval_column(df, date_col, "card_no")

    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="RawData", index=False)

        if "card_no" in df.columns and date_col:
            card_summary = summarize_entities(df, "card_no", date_col=date_col, top_n=top_n_cards, include_intervals=include_intervals)
            if not card_summary.empty:
                if separate_cards:
                    left = card_summary[card_summary["Card Number"].str.startswith("8880")].reset_index(drop=True)
                    right = card_summary[card_summary["Card Number"].str.startswith("8881")].reset_index(drop=True)
                    left.to_excel(writer, sheet_name="TopCards", index=False, startrow=0)
                    startcol = left.shape[1] + 1
                    right.to_excel(writer, sheet_name="TopCards", index=False, startrow=0, startcol=startcol)
                else:
                    card_summary.to_excel(writer, sheet_name="TopCards", index=False)

        if "cashier" in df.columns and date_col:
            cashier_summary = summarize_entities(df, "cashier", date_col=date_col, top_n=top_n_cashiers, include_intervals=include_intervals)
            if not cashier_summary.empty:
                cashier_summary.to_excel(writer, sheet_name="TopCashiers", index=False)

    if separate_cards:
        try:
            wb = load_workbook(output_file)
            if "TopCards" in wb.sheetnames:
                ws = wb["TopCards"]
                max_col = ws.max_column
                blank_col_idx = None
                for c in range(1, max_col + 1):
                    if ws.cell(row=1, column=c).value in (None, ""):
                        blank_col_idx = c
                        break
                if blank_col_idx:
                    left_end = blank_col_idx - 1
                    right_start = blank_col_idx + 1
                else:
                    left_end = max_col // 2
                    right_start = left_end + 2

                blue_fill = PatternFill(start_color="87CEFA", end_color="87CEFA", fill_type="solid")
                yellow_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")

                for col in range(1, left_end + 1):
                    ws.cell(row=1, column=col).fill = blue_fill
                for col in range(right_start, max_col + 1):
                    ws.cell(row=1, column=col).fill = yellow_fill

                wb.save(output_file)
        except Exception:
            pass

def process_file(input_file, top_n_cards=20, top_n_cashiers=20, encrypt=True, separate_cards=False, include_intervals=True):
    script_dir = os.path.dirname(os.path.abspath(__file__))
    output_folder = os.path.join(script_dir, "TopTransactionsPerMonth")
    os.makedirs(output_folder, exist_ok=True)

    password_log_folder = os.path.join(output_folder, "passwordlogs")
    os.makedirs(password_log_folder, exist_ok=True)
    log_file = os.path.join(password_log_folder, "password_log.txt")

    df = pd.read_excel(input_file)
    if "card_no" in df.columns:
        df["card_no"] = df["card_no"].astype(str)

    date_col = _choose_date_col(df)
    if date_col:
        df[date_col] = pd.to_datetime(df[date_col], errors="coerce")
        yearmonths = df[date_col].dt.to_period("M").dropna().unique()
        if len(yearmonths) > 0:
            start = str(min(yearmonths))
            end = str(max(yearmonths))
            month_range = start if start == end else f"{start}_to_{end}"
        else:
            month_range = datetime.now().strftime("%Y-%m")
    else:
        month_range = datetime.now().strftime("%Y-%m")

    output_file = os.path.join(output_folder, f"top_transaction_{month_range}.xlsx")
    process_dynamic_schema(df, output_file, top_n_cards, top_n_cashiers, separate_cards=separate_cards, include_intervals=include_intervals)

    final_file = output_file
    password = None
    if encrypt:
        password = generate_password()
        encrypted_target = output_file.replace(".xlsx", "_encrypted.xlsx")
        try:
            final_file = encrypt_excel(output_file, encrypted_target, password)
            try:
                os.remove(output_file)
            except Exception:
                pass
        except Exception as e:
            raise RuntimeError(f"Failed to encrypt '{output_file}': {e}")

    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    with open(log_file, "a", encoding="utf-8") as log:
        log.write(f"[{timestamp}] Input: {os.path.basename(input_file)} | "
                  f"Output: {os.path.basename(final_file)} | "
                  f"Encryption: {'ENABLED' if encrypt else 'DISABLED'} | "
                  f"Password: {password if encrypt else ''} | Separated: {separate_cards} | "
                  f"IncludeIntervals: {include_intervals}\n")

    print(f"Saved {'and encrypted ' if encrypt else ''}{final_file}")
    return output_folder, final_file, (password if encrypt else None)

def process_entity_details(input_file, card_no=None, cashier=None, include_intervals=True):
    script_dir = os.path.dirname(os.path.abspath(__file__))
    output_folder = os.path.join(script_dir, "TopTransactionsPerMonth", "Card_Cashier_Details_Output")
    os.makedirs(output_folder, exist_ok=True)

    df = pd.read_excel(input_file)
    if "card_no" in df.columns:
        df["card_no"] = df["card_no"].astype(str)

    date_col = _choose_date_col(df)

    if card_no:
        entity_df = df[df["card_no"] == str(card_no)].copy()
        if entity_df.empty:
            raise RuntimeError(f"No rows found for card_no = {card_no}")
        summary_df = summarize_entities(entity_df, "card_no", date_col=date_col, top_n=1, include_intervals=include_intervals)
        safe_card = str(card_no).replace("/", "_").replace("\\", "_")
        output_file = os.path.join(output_folder, f"Card_{safe_card}_details.xlsx")
    elif cashier:
        entity_df = df[df["cashier"] == cashier].copy()
        if entity_df.empty:
            raise RuntimeError(f"No rows found for cashier = {cashier}")
        summary_df = summarize_entities(entity_df, "cashier", date_col=date_col, top_n=1, include_intervals=include_intervals)
        safe_cashier = str(cashier).replace("/", "_").replace("\\", "_")
        output_file = os.path.join(output_folder, f"Cashier_{safe_cashier}_details.xlsx")
    else:
        raise ValueError("Either card_no or cashier must be provided.")

    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        entity_df.to_excel(writer, sheet_name="RawData", index=False)
        if not summary_df.empty:
            summary_df.to_excel(writer, sheet_name="Summary", index=False)

    return output_file
